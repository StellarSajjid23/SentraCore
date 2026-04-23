#!/usr/bin/env python3

import re
import time
import sys
import csv
import os
import ipaddress
from collections import Counter, defaultdict
from datetime import datetime

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

try:
    from colorama import init as colorama_init
    colorama_init()
except ImportError:
    pass


# =========================
# COLORS
# =========================
class Colors:
    RED = "\033[91m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    BLUE = "\033[94m"
    CYAN = "\033[96m"
    WHITE = "\033[97m"
    MAGENTA = "\033[95m"
    BOLD = "\033[1m"
    RESET = "\033[0m"


# =========================
# PATTERNS / TAXONOMY
# =========================
FAILED_PATTERNS = [
    r"failed password",
    r"authentication failed",
    r"login failed",
    r"invalid user",
    r"failed login",
    r"denied",
    r"access denied",
    r"unauthorized"
]

SUCCESS_PATTERNS = [
    r"accepted password",
    r"login successful",
    r"session opened",
    r"successfully logged in",
    r"authentication success",
    r"logged in"
]

SUSPICIOUS_PATTERNS = [
    r"unauthorized",
    r"denied",
    r"error",
    r"warning",
    r"invalid user",
    r"failed password",
    r"root",
    r"admin",
    r"exploit",
    r"attack",
    r"malware",
    r"payload",
    r"blocked",
    r"powershell",
    r"cmd\.exe",
    r"mimikatz",
    r"ransom",
    r"privilege escalation",
    r"brute force",
    r"c2",
    r"beacon",
    r"exfil"
]

EMAIL_PATTERNS = [
    r"subject:",
    r"reply-to",
    r"return-path",
    r"dkim",
    r"spf",
    r"dmarc",
    r"attachment",
    r"mailto:"
]

WEB_PATTERNS = [
    r"http://",
    r"https://",
    r"get /",
    r"post /",
    r"user-agent",
    r"host:",
    r"cookie",
    r"referer",
    r"/api/",
    r"status code",
    r"404",
    r"500"
]

PROCESS_PATTERNS = [
    r"powershell",
    r"cmd\.exe",
    r"bash",
    r"python",
    r"wmic",
    r"regsvr32",
    r"rundll32",
    r"mshta",
    r"certutil",
    r"process",
    r"spawned",
    r"execution"
]

FILE_PATTERNS = [
    r"file",
    r"path",
    r"created",
    r"deleted",
    r"renamed",
    r"write",
    r"\.exe",
    r"\.dll",
    r"\.ps1",
    r"\.js",
    r"\.hta",
    r"/tmp/",
    r"appdata",
    r"programdata"
]

NETWORK_PATTERNS = [
    r"src_ip",
    r"dst_ip",
    r"connection",
    r"port",
    r"tcp",
    r"udp",
    r"icmp",
    r"dns",
    r"firewall",
    r"socket",
    r"flow"
]

TIMESTAMP_PATTERNS = [
    r"\b\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\b",
    r"\b\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\b",
    r"\b[A-Z][a-z]{2}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}\b",
    r"\b\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}(:\d{2})?\b"
]

IP_HEADER_CANDIDATES = {"ip", "ip_address", "source_ip", "src_ip", "client_ip", "remote_ip", "host_ip", "destination_ip", "dst_ip"}
TIMESTAMP_HEADER_CANDIDATES = {"timestamp", "time", "date", "datetime", "event_time", "log_time"}
USERNAME_HEADER_CANDIDATES = {"user", "username", "account", "login", "principal"}
MESSAGE_HEADER_CANDIDATES = {"message", "msg", "event", "description", "details", "log", "entry"}
PROCESS_HEADER_CANDIDATES = {"process", "process_name", "image", "command", "cmdline", "parent_process"}
FILE_HEADER_CANDIDATES = {"file", "file_path", "path", "filename", "object", "target_file"}
URL_HEADER_CANDIDATES = {"url", "uri", "link", "request", "host", "domain"}
STATUS_HEADER_CANDIDATES = {"status", "result", "outcome", "action", "response"}


# =========================
# OUTPUT HELPERS
# =========================
def print_message(message: str):
    print(message + Colors.RESET)


def ask_input(prompt: str) -> str:
    return input(Colors.YELLOW + prompt + Colors.RESET)


# =========================
# BANNER
# =========================
def print_banner():
    banner = r"""
+-----------------------------------------------------------------------------------------------------------------+
|                                                                                                                 |
|      $$\                            $$$$$$\                      $$\                                            |
|      $$ |                          $$  __$$\                     $$ |                                           |
|      $$ |      $$$$$$\   $$$$$$\   $$ /  $$ |$$$$$$$\   $$$$$$\  $$ |$$\   $$\ $$$$$$$$\  $$$$$$\   $$$$$$\     |
|      $$ |     $$  __$$\ $$  __$$\  $$$$$$$$ |$$  __$$\  \____$$\ $$ |$$ |  $$ |\____$$  |$$  __$$\ $$  __$$\    |
|      $$ |     $$ /  $$ |$$ /  $$ | $$  __$$ |$$ |  $$ | $$$$$$$ |$$ |$$ |  $$ |  $$$$ _/ $$$$$$$$ |$$ |  \__|   |
|      $$ |     $$ |  $$ |$$ |  $$ | $$ |  $$ |$$ |  $$ |$$  __$$ |$$ |$$ |  $$ | $$  _/   $$   ____|$$ |         |
|      $$$$$$$$\\$$$$$$  |\$$$$$$$ | $$ |  $$ |$$ |  $$ |\$$$$$$$ |$$ |\$$$$$$$ |$$$$$$$$\ \$$$$$$$\ $$ |         |
|      \________|\______/  \____$$ | \__|  \__|\__|  \__| \_______|\__| \____$$ |\________| \_______|\__|         |
|                         $$\   $$ |                                   $$\   $$ |                                 |
|                         \$$$$$$  |                                   \$$$$$$  |                                 |
|                          \______/                                     \______/                                  |
|                                                                                                                 |
+-----------------------------------------------------------------------------------------------------------------+
|                                  Security Log Triage and Investigation Engine                                   |
+-----------------------------------------------------------------------------------------------------------------+
"""
    print(Colors.RED + banner + Colors.RESET)
    print(Colors.CYAN + "[*] Internship Portfolio Edition" + Colors.RESET)
    print(Colors.GREEN + "[*] Author: StellarSajjid23" + Colors.RESET)
    print(Colors.YELLOW + "[*] Engine: Text / CSV / Excel / Log - SOC Triage Analyzer" + Colors.RESET)
    print("                                                                                               ")


# =========================
# FILE TYPE / HEADER HELPERS
# =========================
def detect_file_type(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()

    if ext in [".txt", ".log"]:
        return "text"
    if ext == ".csv":
        return "csv"
    if ext == ".xlsx":
        return "xlsx"
    if ext == ".xls":
        return "xls"

    return "unsupported"


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def detect_columns(headers: list) -> dict:
    detected = {
        "timestamp": None,
        "ip": None,
        "username": None,
        "message": None,
        "process": None,
        "file": None,
        "url": None,
        "status": None
    }

    for idx, header in enumerate(headers):
        if header in TIMESTAMP_HEADER_CANDIDATES and detected["timestamp"] is None:
            detected["timestamp"] = idx
        if header in IP_HEADER_CANDIDATES and detected["ip"] is None:
            detected["ip"] = idx
        if header in USERNAME_HEADER_CANDIDATES and detected["username"] is None:
            detected["username"] = idx
        if header in MESSAGE_HEADER_CANDIDATES and detected["message"] is None:
            detected["message"] = idx
        if header in PROCESS_HEADER_CANDIDATES and detected["process"] is None:
            detected["process"] = idx
        if header in FILE_HEADER_CANDIDATES and detected["file"] is None:
            detected["file"] = idx
        if header in URL_HEADER_CANDIDATES and detected["url"] is None:
            detected["url"] = idx
        if header in STATUS_HEADER_CANDIDATES and detected["status"] is None:
            detected["status"] = idx

    return detected


def row_to_log_entry(row: list, columns: dict) -> tuple[str, dict]:
    normalized = {}
    parts = []

    for field_name in ["timestamp", "ip", "username", "process", "file", "url", "status", "message"]:
        idx = columns.get(field_name)
        value = ""
        if idx is not None and idx < len(row):
            value = str(row[idx]).strip()
        normalized[field_name] = value
        if value:
            parts.append(value)

    if not parts:
        parts = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]

    return " | ".join(parts), normalized


# =========================
# LOADERS
# =========================
def load_text_file(file_path: str):
    lines = []
    rows = []
    with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
        for line in file:
            cleaned = line.strip()
            if cleaned:
                lines.append(cleaned)
                rows.append({
                    "timestamp": "",
                    "ip": "",
                    "username": "",
                    "process": "",
                    "file": "",
                    "url": "",
                    "status": "",
                    "message": cleaned
                })

    empty_cols = {"timestamp": None, "ip": None, "username": None, "message": None, "process": None, "file": None, "url": None, "status": None}
    return lines, empty_cols, rows, 0


def load_csv_file(file_path: str):
    rows = []
    normalized_rows = []
    malformed_rows = 0

    with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as file:
        reader = csv.reader(file)
        all_rows = list(reader)

    if not all_rows:
        empty_cols = {"timestamp": None, "ip": None, "username": None, "message": None, "process": None, "file": None, "url": None, "status": None}
        return [], empty_cols, [], 0

    headers = [normalize_header(h) for h in all_rows[0]]
    columns = detect_columns(headers)
    has_detected_headers = any(v is not None for v in columns.values())
    data_rows = all_rows[1:] if has_detected_headers else all_rows
    expected_len = len(all_rows[0])

    for row in data_rows:
        if not row:
            continue
        if len(row) != expected_len:
            malformed_rows += 1
        entry, normalized = row_to_log_entry(row, columns)
        if entry:
            rows.append(entry)
            normalized_rows.append(normalized)

    return rows, columns, normalized_rows, malformed_rows


def load_xlsx_file(file_path: str):
    if not OPENPYXL_AVAILABLE:
        print_message(Colors.RED + "[!] openpyxl is required to read .xlsx files.")
        print_message(Colors.YELLOW + "Install it with: py -m pip install openpyxl")
        sys.exit(1)

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    rows = []
    normalized_rows = []
    columns = {"timestamp": None, "ip": None, "username": None, "message": None, "process": None, "file": None, "url": None, "status": None}
    header_detected = False
    malformed_rows = 0

    for sheet in workbook.worksheets:
        first_row = True
        sheet_columns = columns.copy()
        expected_len = None

        for row in sheet.iter_rows(values_only=True):
            cleaned_row = ["" if cell is None else str(cell).strip() for cell in row]

            if not any(cleaned_row):
                continue

            if expected_len is None:
                expected_len = len(cleaned_row)
            elif len(cleaned_row) != expected_len:
                malformed_rows += 1

            if first_row:
                normalized_headers = [normalize_header(cell) for cell in cleaned_row]
                detected = detect_columns(normalized_headers)

                if any(v is not None for v in detected.values()):
                    sheet_columns = detected
                    columns = detected
                    header_detected = True
                    first_row = False
                    continue

                first_row = False

            entry, normalized = row_to_log_entry(cleaned_row, sheet_columns)
            if entry:
                rows.append(entry)
                normalized_rows.append(normalized)

    if not header_detected:
        columns = {"timestamp": None, "ip": None, "username": None, "message": None, "process": None, "file": None, "url": None, "status": None}

    return rows, columns, normalized_rows, malformed_rows


def load_xls_file(file_path: str):
    if not XLRD_AVAILABLE:
        print_message(Colors.RED + "[!] xlrd is required to read .xls files.")
        print_message(Colors.YELLOW + "Install it with: py -m pip install xlrd")
        sys.exit(1)

    workbook = xlrd.open_workbook(file_path)
    rows = []
    normalized_rows = []
    columns = {"timestamp": None, "ip": None, "username": None, "message": None, "process": None, "file": None, "url": None, "status": None}
    header_detected = False
    malformed_rows = 0

    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        first_row = True
        sheet_columns = columns.copy()
        expected_len = None

        for row_idx in range(sheet.nrows):
            row = [str(sheet.cell_value(row_idx, col_idx)).strip() for col_idx in range(sheet.ncols)]

            if not any(row):
                continue

            if expected_len is None:
                expected_len = len(row)
            elif len(row) != expected_len:
                malformed_rows += 1

            if first_row:
                normalized_headers = [normalize_header(cell) for cell in row]
                detected = detect_columns(normalized_headers)

                if any(v is not None for v in detected.values()):
                    sheet_columns = detected
                    columns = detected
                    header_detected = True
                    first_row = False
                    continue

                first_row = False

            entry, normalized = row_to_log_entry(row, sheet_columns)
            if entry:
                rows.append(entry)
                normalized_rows.append(normalized)

    if not header_detected:
        columns = {"timestamp": None, "ip": None, "username": None, "message": None, "process": None, "file": None, "url": None, "status": None}

    return rows, columns, normalized_rows, malformed_rows


def load_log_file(file_path: str):
    if not os.path.exists(file_path):
        print_message(Colors.RED + f"[!] File not found: {file_path}")
        sys.exit(1)

    file_type = detect_file_type(file_path)

    try:
        if file_type == "text":
            return load_text_file(file_path)
        if file_type == "csv":
            return load_csv_file(file_path)
        if file_type == "xlsx":
            return load_xlsx_file(file_path)
        if file_type == "xls":
            return load_xls_file(file_path)

        print_message(Colors.RED + "[!] Unsupported file type.")
        print_message(Colors.YELLOW + "Supported types: .txt, .log, .csv, .xlsx, .xls")
        sys.exit(1)
    except Exception as exc:
        print_message(Colors.RED + f"[!] Failed to read log file: {exc}")
        sys.exit(1)


# =========================
# ENRICHMENT / EXTRACTION
# =========================
def extract_ip_addresses(line: str) -> list:
    return re.findall(r"\b(?:\d{1,3}\.){3}\d{1,3}\b", line)


def is_public_ip(ip: str) -> bool:
    try:
        ip_obj = ipaddress.ip_address(ip)
        return ip_obj.is_global
    except ValueError:
        return False


def get_ip_geo(ip: str) -> dict:
    result = {
        "country": "Unknown",
        "city": "Unknown",
        "isp": "Unknown",
        "note": "Unavailable"
    }

    if not is_public_ip(ip):
        result["note"] = "Private/Internal"
        return result

    if not REQUESTS_AVAILABLE:
        result["note"] = "requests missing"
        return result

    try:
        response = requests.get(f"http://ip-api.com/json/{ip}", timeout=4)
        data = response.json()
        if data.get("status") == "success":
            result["country"] = data.get("country", "Unknown")
            result["city"] = data.get("city", "Unknown")
            result["isp"] = data.get("isp", "Unknown")
            result["note"] = "Success"
        else:
            result["note"] = "Lookup Failed"
    except Exception:
        result["note"] = "Lookup Failed"

    return result


def detect_timestamp(line: str, normalized: dict) -> str:
    if normalized.get("timestamp"):
        return normalized["timestamp"]

    for pattern in TIMESTAMP_PATTERNS:
        match = re.search(pattern, line)
        if match:
            return match.group(0)

    return "Unknown"


def classify_event_category(line: str, normalized: dict) -> str:
    lowered = line.lower()

    if normalized.get("url") or any(re.search(pattern, lowered) for pattern in WEB_PATTERNS):
        return "Web"
    if normalized.get("file") or any(re.search(pattern, lowered) for pattern in FILE_PATTERNS):
        return "File"
    if normalized.get("process") or any(re.search(pattern, lowered) for pattern in PROCESS_PATTERNS):
        return "Process"
    if normalized.get("username") or any(re.search(pattern, lowered) for pattern in FAILED_PATTERNS + SUCCESS_PATTERNS):
        return "Authentication"
    if normalized.get("ip") or any(re.search(pattern, lowered) for pattern in NETWORK_PATTERNS):
        return "Network"
    if any(re.search(pattern, lowered) for pattern in EMAIL_PATTERNS):
        return "Email"
    return "System"


def classify_severity(line: str, matched_failed: bool, matched_success: bool, matched_suspicious: bool, category: str) -> str:
    lowered = line.lower()

    high_markers = [
        "unauthorized", "exploit", "attack", "malware", "payload", "blocked",
        "privilege escalation", "mimikatz", "ransom", "root login", "admin abuse"
    ]
    medium_markers = [
        "warning", "invalid user", "failed login", "failed password", "error",
        "denied", "access denied", "powershell", "cmd.exe"
    ]

    if any(marker in lowered for marker in high_markers):
        return "High"

    if matched_failed or any(marker in lowered for marker in medium_markers):
        return "Medium"

    if category in {"Process", "File"} and matched_suspicious:
        return "Medium"

    if matched_success or matched_suspicious:
        return "Low"

    return "Low"


def determine_triage_priority(severity: str, category: str, public_ips: list) -> str:
    if severity == "High" and public_ips:
        return "Urgent"
    if severity == "High":
        return "High"
    if severity == "Medium" and category in {"Authentication", "Network", "Process"}:
        return "Medium"
    return "Low"


def build_event_story(category: str, severity: str, public_ips: list, username: str, process: str, file_value: str) -> str:
    parts = [category, severity]

    if public_ips:
        parts.append(f"External IPs={len(public_ips)}")
    if username:
        parts.append(f"User={username[:20]}")
    if process:
        parts.append(f"Process={process[:20]}")
    if file_value:
        parts.append(f"File={file_value[:20]}")

    return " | ".join(parts)


# =========================
# MAIN ANALYSIS
# =========================
def analyze_log_lines(lines: list, normalized_rows: list, malformed_rows: int) -> dict:
    failed_count = 0
    success_count = 0
    suspicious_count = 0

    failed_lines = []
    success_lines = []
    suspicious_lines = []
    analyst_findings = []

    severity_counter = Counter()
    priority_counter = Counter()
    category_counter = Counter()
    missing_field_counter = Counter()

    failed_ip_counter = Counter()
    suspicious_ip_counter = Counter()
    public_ip_counter = Counter()
    country_counter = Counter()
    username_counter = Counter()
    process_counter = Counter()
    file_counter = Counter()
    url_counter = Counter()
    time_bucket_counter = Counter()

    event_group_counter = Counter()

    total_rows = len(normalized_rows)
    rows_with_detected_fields = 0

    for idx, line in enumerate(lines):
        normalized = normalized_rows[idx] if idx < len(normalized_rows) else {}

        if any(normalized.get(field, "") for field in normalized):
            rows_with_detected_fields += 1

        for field_name, value in normalized.items():
            if value == "":
                missing_field_counter[field_name] += 1

        category = classify_event_category(line, normalized)
        category_counter[category] += 1

        lowered = line.lower().strip()
        ips = extract_ip_addresses(line)
        public_ips = [ip for ip in ips if is_public_ip(ip)]

        username = normalized.get("username", "")
        process_value = normalized.get("process", "")
        file_value = normalized.get("file", "")
        url_value = normalized.get("url", "")
        timestamp = detect_timestamp(line, normalized)

        matched_failed = any(re.search(pattern, lowered) for pattern in FAILED_PATTERNS)
        matched_success = any(re.search(pattern, lowered) for pattern in SUCCESS_PATTERNS)
        matched_suspicious = any(re.search(pattern, lowered) for pattern in SUSPICIOUS_PATTERNS)

        severity = classify_severity(line, matched_failed, matched_success, matched_suspicious, category)
        priority = determine_triage_priority(severity, category, public_ips)
        severity_counter[severity] += 1
        priority_counter[priority] += 1

        if username:
            username_counter[username] += 1
        if process_value:
            process_counter[process_value] += 1
        if file_value:
            file_counter[file_value] += 1
        if url_value:
            url_counter[url_value] += 1

        if timestamp != "Unknown":
            time_bucket_counter[timestamp[:13]] += 1

        if matched_failed:
            failed_count += 1
            failed_lines.append((line.strip(), severity, category, priority))
            for ip in ips:
                failed_ip_counter[ip] += 1

        if matched_success:
            success_count += 1
            success_lines.append((line.strip(), severity, category, priority))

        if matched_suspicious:
            suspicious_count += 1
            suspicious_lines.append((line.strip(), severity, category, priority))
            for ip in ips:
                suspicious_ip_counter[ip] += 1

        for ip in public_ips:
            public_ip_counter[ip] += 1

        for ip in set(public_ips):
            geo = get_ip_geo(ip)
            country_counter[geo["country"]] += 1

        event_signature = (
            category,
            severity,
            username[:20],
            process_value[:20],
            file_value[:20],
            tuple(sorted(set(public_ips)))
        )
        event_group_counter[event_signature] += 1

        if severity == "High" or priority in {"Urgent", "High"}:
            analyst_findings.append({
                "timestamp": timestamp,
                "category": category,
                "severity": severity,
                "priority": priority,
                "line": line.strip(),
                "story": build_event_story(category, severity, public_ips, username, process_value, file_value)
            })

    brute_force_candidates = []
    for ip, count in failed_ip_counter.items():
        if count >= 5:
            brute_force_candidates.append((ip, count))
    brute_force_candidates.sort(key=lambda x: x[1], reverse=True)

    repeated_event_clusters = []
    for signature, count in event_group_counter.items():
        if count >= 3:
            category, severity, username, process_value, file_value, ip_tuple = signature
            repeated_event_clusters.append({
                "count": count,
                "category": category,
                "severity": severity,
                "username": username or "-",
                "process": process_value or "-",
                "file": file_value or "-",
                "ips": ",".join(ip_tuple[:2]) if ip_tuple else "-"
            })
    repeated_event_clusters.sort(key=lambda x: x["count"], reverse=True)

    analyst_findings.sort(
        key=lambda item: (
            {"Urgent": 4, "High": 3, "Medium": 2, "Low": 1}.get(item["priority"], 0),
            {"High": 3, "Medium": 2, "Low": 1}.get(item["severity"], 0)
        ),
        reverse=True
    )

    field_quality_score = 0
    if total_rows > 0:
        completeness = 1 - (sum(missing_field_counter.values()) / max(total_rows * 8, 1))
        field_quality_score = round(max(completeness, 0) * 100, 2)

    normalization_score = 0
    if total_rows > 0:
        normalization_score = round((rows_with_detected_fields / total_rows) * 100, 2)

    return {
        "total_lines": len(lines),
        "failed_count": failed_count,
        "success_count": success_count,
        "suspicious_count": suspicious_count,
        "failed_lines": failed_lines,
        "success_lines": success_lines,
        "suspicious_lines": suspicious_lines,
        "top_failed_ips": failed_ip_counter.most_common(10),
        "top_suspicious_ips": suspicious_ip_counter.most_common(10),
        "top_public_ips": public_ip_counter.most_common(10),
        "top_usernames": username_counter.most_common(10),
        "top_processes": process_counter.most_common(10),
        "top_files": file_counter.most_common(10),
        "top_urls": url_counter.most_common(10),
        "brute_force_candidates": brute_force_candidates,
        "severity_counter": severity_counter,
        "priority_counter": priority_counter,
        "category_counter": category_counter,
        "missing_field_counter": missing_field_counter,
        "malformed_rows": malformed_rows,
        "field_quality_score": field_quality_score,
        "normalization_score": normalization_score,
        "country_counter": country_counter,
        "time_bucket_counter": time_bucket_counter,
        "repeated_event_clusters": repeated_event_clusters,
        "analyst_findings": analyst_findings
    }


# =========================
# RENDERING
# =========================
def render_detected_columns(columns: dict):
    print("\n" + Colors.CYAN + Colors.BOLD + "Detected Columns:" + Colors.RESET)

    border = "+------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'Field':^25}|{'Detected Column Index':^28}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    for field in ["timestamp", "ip", "username", "message", "process", "file", "url", "status"]:
        value = columns.get(field)
        value_text = str(value) if value is not None else "Not Detected"
        color = Colors.GREEN if value is not None else Colors.YELLOW

        print(
            Colors.WHITE + "|" +
            f"{field.capitalize():<25}" +
            "|" +
            color + f"{value_text:^28}" +
            Colors.WHITE + "|" +
            Colors.RESET
        )

    print(Colors.CYAN + border + Colors.RESET)


def render_summary(result: dict):
    print("\n" + Colors.CYAN + Colors.BOLD + "Log Analysis Summary:" + Colors.RESET)

    border = "+------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'Checking':^38}|{'Status':^15}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    rows = [
        ("Total Log Lines", str(result["total_lines"]), Colors.WHITE),
        ("Failed Logins", str(result["failed_count"]), Colors.RED),
        ("Successful Logins", str(result["success_count"]), Colors.GREEN),
        ("Suspicious Events", str(result["suspicious_count"]), Colors.YELLOW),
        ("Brute Force IPs", str(len(result["brute_force_candidates"])), Colors.MAGENTA),
        ("Malformed Rows", str(result["malformed_rows"]), Colors.CYAN),
        ("Field Quality Score", f"{result['field_quality_score']}%", Colors.BLUE),
        ("Normalization Score", f"{result['normalization_score']}%", Colors.BLUE),
    ]

    for label, value, color in rows:
        print(
            Colors.WHITE + "|" +
            f"{label:<38}" +
            "|" +
            color + f"{value:^15}" +
            Colors.WHITE + "|" +
            Colors.RESET
        )

    print(Colors.CYAN + border + Colors.RESET)


def render_distribution_table(title: str, counter: Counter, primary_color: str = None):
    print("\n" + Colors.MAGENTA + Colors.BOLD + title + ":" + Colors.RESET)

    border = "+------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'Value':^38}|{'Count':^15}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    items = counter.most_common()
    if not items:
        print(Colors.WHITE + f"|{'None':^38}|{'0':^15}|" + Colors.RESET)
    else:
        for label, count in items:
            color = primary_color or Colors.YELLOW
            if title == "Severity Distribution":
                color = Colors.GREEN
                if label == "High":
                    color = Colors.RED
                elif label == "Medium":
                    color = Colors.YELLOW
            if title == "Priority Distribution":
                color = Colors.GREEN
                if label == "Urgent":
                    color = Colors.RED
                elif label == "High":
                    color = Colors.MAGENTA
                elif label == "Medium":
                    color = Colors.YELLOW

            print(
                Colors.WHITE + "|" +
                f"{str(label)[:38]:<38}" +
                "|" +
                color + f"{str(count):^15}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_ip_table(title: str, items: list, color: str):
    print("\n" + color + Colors.BOLD + title + ":" + Colors.RESET)

    border = "+------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'IP Address':^38}|{'Count':^15}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    if not items:
        print(Colors.WHITE + f"|{'None':^38}|{'0':^15}|" + Colors.RESET)
    else:
        for ip, count in items:
            print(
                Colors.WHITE + "|" +
                color + f"{ip[:38]:<38}" +
                Colors.WHITE + "|" +
                color + f"{str(count):^15}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_entity_table(title: str, items: list, color: str):
    print("\n" + color + Colors.BOLD + title + ":" + Colors.RESET)

    border = "+------------------------------------------------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'Value':^80}|{'Count':^15}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    if not items:
        print(Colors.WHITE + f"|{'None':^80}|{'0':^15}|" + Colors.RESET)
    else:
        for value, count in items:
            print(
                Colors.WHITE + "|" +
                color + f"{str(value)[:80]:<80}" +
                Colors.WHITE + "|" +
                color + f"{str(count):^15}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_missing_fields_table(counter: Counter):
    print("\n" + Colors.MAGENTA + Colors.BOLD + "Missing Field Summary:" + Colors.RESET)

    border = "+------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'Field':^38}|{'Missing Count':^15}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    items = counter.most_common()
    if not items:
        print(Colors.WHITE + f"|{'None':^38}|{'0':^15}|" + Colors.RESET)
    else:
        for label, count in items:
            print(
                Colors.WHITE + "|" +
                f"{label:<38}" +
                "|" +
                Colors.YELLOW + f"{str(count):^15}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_line_table(title: str, items: list, color: str, limit: int = 10):
    print("\n" + color + Colors.BOLD + title + ":" + Colors.RESET)

    border = "+-----------------------------------------------------------------------------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'#':^5}|{'Severity':^10}|{'Category':^16}|{'Priority':^10}|{'Log Entry':^80}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    if not items:
        print(Colors.WHITE + f"|{'-':^5}|{'-':^10}|{'-':^16}|{'-':^10}|{'None':^80}|" + Colors.RESET)
    else:
        for idx, item in enumerate(items[:limit], start=1):
            line, severity, category, priority = item
            shortened = line[:80]

            sev_color = Colors.GREEN
            if severity == "High":
                sev_color = Colors.RED
            elif severity == "Medium":
                sev_color = Colors.YELLOW

            pri_color = Colors.GREEN
            if priority == "Urgent":
                pri_color = Colors.RED
            elif priority == "High":
                pri_color = Colors.MAGENTA
            elif priority == "Medium":
                pri_color = Colors.YELLOW

            print(
                Colors.WHITE + "|" +
                f"{str(idx):^5}" +
                "|" +
                sev_color + f"{severity:^10}" +
                Colors.WHITE + "|" +
                Colors.CYAN + f"{category:^16}" +
                Colors.WHITE + "|" +
                pri_color + f"{priority:^10}" +
                Colors.WHITE + "|" +
                color + f"{shortened:<80}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_cluster_table(clusters: list, limit: int = 10):
    print("\n" + Colors.CYAN + Colors.BOLD + "Repeated Event Clusters:" + Colors.RESET)

    border = "+-------------------------------------------------------------------------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'#':^5}|{'Count':^8}|{'Category':^14}|{'Severity':^10}|{'User':^18}|{'Process':^18}|{'File/IP Hint':^42}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    if not clusters:
        print(Colors.WHITE + f"|{'-':^5}|{'0':^8}|{'None':^14}|{'-':^10}|{'-':^18}|{'-':^18}|{'None':^42}|" + Colors.RESET)
    else:
        for idx, item in enumerate(clusters[:limit], start=1):
            sev_color = Colors.GREEN
            if item["severity"] == "High":
                sev_color = Colors.RED
            elif item["severity"] == "Medium":
                sev_color = Colors.YELLOW

            file_ip_hint = f"{item['file']} | {item['ips']}"[:42]

            print(
                Colors.WHITE + "|" +
                f"{idx:^5}" +
                "|" +
                Colors.MAGENTA + f"{str(item['count']):^8}" +
                Colors.WHITE + "|" +
                Colors.CYAN + f"{item['category'][:14]:^14}" +
                Colors.WHITE + "|" +
                sev_color + f"{item['severity']:^10}" +
                Colors.WHITE + "|" +
                f"{item['username'][:18]:^18}" +
                "|" +
                f"{item['process'][:18]:^18}" +
                "|" +
                Colors.YELLOW + f"{file_ip_hint:<42}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_geo_table(items: list, limit: int = 8):
    print("\n" + Colors.CYAN + Colors.BOLD + "Top Public IP Geo Enrichment:" + Colors.RESET)

    border = "+----------------------------------------------------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'#':^5}|{'IP':^18}|{'Country':^18}|{'City':^18}|{'ISP':^28}|{'Hits':^8}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    if not items:
        print(Colors.WHITE + f"|{'-':^5}|{'None':^18}|{'None':^18}|{'None':^18}|{'None':^28}|{'0':^8}|" + Colors.RESET)
    else:
        for idx, (ip, count) in enumerate(items[:limit], start=1):
            geo = get_ip_geo(ip)
            print(
                Colors.WHITE + "|" +
                f"{idx:^5}" +
                "|" +
                Colors.YELLOW + f"{ip[:18]:^18}" +
                Colors.WHITE + "|" +
                f"{geo['country'][:18]:^18}" +
                "|" +
                f"{geo['city'][:18]:^18}" +
                "|" +
                f"{geo['isp'][:28]:^28}" +
                "|" +
                Colors.MAGENTA + f"{str(count):^8}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_analyst_findings(findings: list, limit: int = 10):
    print("\n" + Colors.RED + Colors.BOLD + "Analyst Priority Findings:" + Colors.RESET)

    border = "+------------------------------------------------------------------------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'#':^5}|{'Priority':^10}|{'Severity':^10}|{'Category':^15}|{'Time':^20}|{'Storyline':^55}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    if not findings:
        print(Colors.WHITE + f"|{'-':^5}|{'-':^10}|{'-':^10}|{'None':^15}|{'None':^20}|{'No elevated findings':^55}|" + Colors.RESET)
    else:
        for idx, finding in enumerate(findings[:limit], start=1):
            pri_color = Colors.GREEN
            if finding["priority"] == "Urgent":
                pri_color = Colors.RED
            elif finding["priority"] == "High":
                pri_color = Colors.MAGENTA
            elif finding["priority"] == "Medium":
                pri_color = Colors.YELLOW

            sev_color = Colors.GREEN
            if finding["severity"] == "High":
                sev_color = Colors.RED
            elif finding["severity"] == "Medium":
                sev_color = Colors.YELLOW

            print(
                Colors.WHITE + "|" +
                f"{idx:^5}" +
                "|" +
                pri_color + f"{finding['priority']:^10}" +
                Colors.WHITE + "|" +
                sev_color + f"{finding['severity']:^10}" +
                Colors.WHITE + "|" +
                Colors.CYAN + f"{finding['category'][:15]:^15}" +
                Colors.WHITE + "|" +
                f"{finding['timestamp'][:20]:^20}" +
                "|" +
                Colors.YELLOW + f"{finding['story'][:55]:<55}" +
                Colors.WHITE + "|" +
                Colors.RESET
            )

    print(Colors.CYAN + border + Colors.RESET)


def render_recommendations(result: dict):
    print("\n" + Colors.CYAN + Colors.BOLD + "Recommendations:" + Colors.RESET)

    border = "+-----------------------------------------------------------------------------+"
    print(Colors.CYAN + border + Colors.RESET)
    print(Colors.WHITE + f"|{'#':^6}|{'Recommendation':^70}|" + Colors.RESET)
    print(Colors.CYAN + border + Colors.RESET)

    recommendations = []

    if result["priority_counter"].get("Urgent", 0) > 0:
        recommendations.append("Investigate urgent findings first.")
        recommendations.append("Correlate public IP activity with firewall or proxy logs.")
    elif result["priority_counter"].get("High", 0) > 0:
        recommendations.append("Review high-priority authentication/network/process events.")
    else:
        recommendations.append("No critical findings detected; continue routine monitoring.")

    if result["brute_force_candidates"]:
        recommendations.append("Check repeated failed login IPs for brute-force behavior.")
    if result["top_public_ips"]:
        recommendations.append("Review external IP countries/ISPs for unusual access patterns.")
    if result["top_processes"]:
        recommendations.append("Inspect suspicious process executions for LOLBins or abuse.")
    if result["top_files"]:
        recommendations.append("Validate repeated file paths for persistence or malware activity.")
    if result["repeated_event_clusters"]:
        recommendations.append("Clustered repeated events may indicate automation or attack tooling.")
    if result["field_quality_score"] < 50:
        recommendations.append("Improve log field consistency to raise triage accuracy.")

    deduped = []
    for item in recommendations:
        if item not in deduped:
            deduped.append(item)

    for idx, item in enumerate(deduped[:8], start=1):
        text = item[:56]
        print(
            Colors.WHITE + "|" +
            f"{str(idx):^6}" +
            "|" +
            Colors.YELLOW + f"{text:<70}" +
            Colors.WHITE + "|" +
            Colors.RESET
        )

    print(Colors.CYAN + border + Colors.RESET)


# =========================
# MAIN
# =========================
def main():
    print_banner()

    print_message(Colors.BLUE + "[i] Mode        : Security Log Triage")
    print_message(Colors.BLUE + "[i] Input Type  : Text / CSV / Excel")
    print_message(Colors.BLUE + "[i] Detection   : Failed Login / Suspicious Activity / Event Taxonomy")
    print_message(Colors.BLUE + "[i] Features    : Schema Detection + Priority Triage + GeoIP + Clustering\n")

    try:
        log_file = ask_input("Enter File Path to [ .txt / .log / .csv / .xlsx / .xls ] : ").strip()

        if not log_file:
            print_message(Colors.RED + "[!] No file path entered.")
            sys.exit(1)

        print()
        print_message(Colors.YELLOW + "[-] Loading File ...")
        lines, columns, normalized_rows, malformed_rows = load_log_file(log_file)

        print_message(Colors.YELLOW + "[-] Analyzing Entries ...")
        print_message(Colors.YELLOW + "[-] Building Triage Views and Entity Pivots ...\n")

        result = analyze_log_lines(lines, normalized_rows, malformed_rows)

        render_detected_columns(columns)
        render_summary(result)
        render_distribution_table("Severity Distribution", result["severity_counter"])
        render_distribution_table("Priority Distribution", result["priority_counter"])
        render_distribution_table("Event Category Distribution", result["category_counter"])
        render_distribution_table("Country Distribution", result["country_counter"], Colors.YELLOW)
        render_missing_fields_table(result["missing_field_counter"])
        render_ip_table("Top Failed Login IPs", result["top_failed_ips"], Colors.RED)
        render_ip_table("Top Suspicious IPs", result["top_suspicious_ips"], Colors.MAGENTA)
        render_ip_table("Brute Force Candidates", result["brute_force_candidates"], Colors.YELLOW)
        render_geo_table(result["top_public_ips"])
        render_entity_table("Top Usernames", result["top_usernames"], Colors.CYAN)
        render_entity_table("Top Processes", result["top_processes"], Colors.YELLOW)
        render_entity_table("Top Files", result["top_files"], Colors.GREEN)
        render_entity_table("Top URLs", result["top_urls"], Colors.MAGENTA)
        render_cluster_table(result["repeated_event_clusters"])
        render_analyst_findings(result["analyst_findings"])
        render_line_table("Sample Failed Login Events", result["failed_lines"], Colors.RED)
        render_line_table("Sample Suspicious Events", result["suspicious_lines"], Colors.YELLOW)
        render_line_table("Sample Successful Login Events", result["success_lines"], Colors.GREEN)
        render_recommendations(result)

    except KeyboardInterrupt:
        print_message("\n" + Colors.RED + "[!] Analysis interrupted by user.")
        sys.exit(0)
    except Exception as exc:
        print_message(Colors.RED + f"[!] Unexpected error: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    main()
    time.sleep(60)
